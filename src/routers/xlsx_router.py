import asyncio
import html  # для декодирования HTML-кодов
import logging
import os
import tempfile
import time
from datetime import datetime
from typing import List, Dict

import requests
from fastapi import APIRouter, HTTPException
from fastapi.responses import JSONResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pydantic import BaseModel

from config import settings

router = APIRouter()

MEGAPLAN_API_URL = settings.MEGAPLAN_API_URL
MEGAPLAN_API_KEY = settings.MEGAPLAN_API_KEY
MEGAPLAN_HEADER = {
    "Authorization": f"Bearer {MEGAPLAN_API_KEY}",
    "Content-Type": "application/json"
}

# Словарь с русскими названиями месяцев
MONTHS_RU = {
    1: 'января',
    2: 'февраля',
    3: 'марта',
    4: 'апреля',
    5: 'мая',
    6: 'июня',
    7: 'июля',
    8: 'августа',
    9: 'сентября',
    10: 'октября',
    11: 'ноября',
    12: 'декабря'
}


def get_project_issues(project_id: str, url: str, header: Dict[str, str]) -> List[Dict]:
    url = f"{url}/api/v3/project/{project_id}/issues"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        project_data = response.json()["data"]
        logging.info(f"Получены задачи проекта с ID: {project_id}")
        time.sleep(1)
        return project_data
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting project {project_id}: {e}")
        raise


def get_project(project_id: str, url: str, header: Dict[str, str]) -> Dict:
    url = f"{url}/api/v3/project/{project_id}"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        project_data = response.json()["data"]
        logging.info(f"Получен проект с ID: {project_id}")
        time.sleep(1)
        return project_data
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting project {project_id}: {e}")
        raise


def get_task(task_id: str, url: str, header: Dict[str, str]) -> Dict:
    url = f"{url}/api/v3/task/{task_id}"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        task_data = response.json()["data"]
        logging.info(f"Получена задача с ID: {task_id}")
        time.sleep(1)
        return task_data
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting task {task_id}: {e}")
        raise


def get_task_subtasks(task_id: str, url: str, header: Dict[str, str]) -> List[Dict]:
    url = f"{url}/api/v3/task/{task_id}/subTasks"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        subtasks = response.json()["data"]
        logging.info(f"Получены подзадачи задачи с ID: {task_id}")
        time.sleep(1)
        return subtasks
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting subtasks for task {task_id}: {e}")
        raise


def get_comment(comment_id: str, url: str, header: Dict[str, str]) -> str:
    url = f"{url}/api/v3/comment/{comment_id}"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        comment_data = response.json()["data"]
        logging.info(f"Получен комментарий с ID: {comment_id}")
        time.sleep(1)
        clean_comment = clean_html(comment_data["content"])
        return clean_comment
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting comment {comment_id}: {e}")
        return ""


def get_employee(employee_id, url, header):
    url = f"{url}/api/v3/employee/{employee_id}"
    try:
        response = requests.get(url, headers=header, timeout=120)
        response.raise_for_status()
        employee_data = response.json()["data"]
        logging.info(f"Получен сотрудник с ID: {employee_id}")
        time.sleep(1)
        return employee_data
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error occurred while getting employee {employee_id}: {e}")
        raise


def get_responsible_name(responsible_data, url, header):
    if "name" in responsible_data:
        return responsible_data["name"]
    else:
        employee_id = responsible_data["id"]
        employee_data = get_employee(employee_id, url, header)
        return employee_data["name"]


def is_product(text: str) -> bool:
    return any(char.isdigit() for char in text)


def clean_html(text: str) -> str:
    """Удаление HTML-тегов и декодирование символов."""
    text = html.unescape(text)  # Декодируем HTML-символы
    text = text.replace('<br />', '\n').replace('</p>', '\n').replace('<p>', '').replace('</strong>', '').replace(
        '<strong>', '')
    return text.strip()


def process_tasks(project_name: str, issues: List[Dict], sheet, project_responsible) -> None:
    row = 2
    for issue in issues:
        issue_name = issue["name"]
        issue_data = get_task(issue["id"], MEGAPLAN_API_URL, MEGAPLAN_HEADER)

        development_task = next(
            (task for task in issue_data["subTasks"] if "разработка продуктов" in task["name"].lower()), None)
        if development_task:
            logging.info(f'Получена задача {development_task["name"]} с ID {development_task["id"]}')
            development_task_data = get_task(development_task["id"], MEGAPLAN_API_URL, MEGAPLAN_HEADER)
            owner_name = get_responsible_name(development_task_data["responsible"], MEGAPLAN_API_URL, MEGAPLAN_HEADER)

            # Получение комментариев и имени ответственного перед циклом
            raw_materials_comment = ""
            packaging_comment = ""
            last_comment = ""

            raw_materials_task = next(
                (task for task in development_task_data["subTasks"] if task["name"] == "1. Поставщики сырья"), None)
            if raw_materials_task:
                logging.info(f'Получена задача 1. Поставщики сырья с ID {raw_materials_task["id"]}')
                raw_materials_task_data = get_task(raw_materials_task["id"], MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                if raw_materials_task_data["comments"]:
                    raw_materials_comment = get_comment(raw_materials_task_data["comments"][0]["id"], MEGAPLAN_API_URL,
                                                        MEGAPLAN_HEADER)

            packaging_task = next(
                (task for task in development_task_data["subTasks"] if task["name"] == "2. Поставщики упаковки"), None)
            if packaging_task:
                logging.info(f'Получена задача 2. Поставщики упаковки с ID {packaging_task["id"]}')
                packaging_task_data = get_task(packaging_task["id"], MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                if packaging_task_data["comments"]:
                    packaging_comment = get_comment(packaging_task_data["comments"][0]["id"], MEGAPLAN_API_URL,
                                                    MEGAPLAN_HEADER)

            if development_task_data["comments"]:
                last_comment = get_comment(development_task_data["comments"][-1]["id"], MEGAPLAN_API_URL,
                                           MEGAPLAN_HEADER)

            # Декодирование и очистка данных
            products_raw = development_task_data["subject"]
            products_clean = clean_html(products_raw)
            logging.info(f"Продукты: {products_clean}")

            products = products_clean.split("\n\n")  # Разделение продуктов
            if len(products) == 1:
                if all(el[0].isdigit() for el in products_clean.split("\n") if el):
                    products = products_clean.split("\n")

            # Форматируем дату
            raw_date = issue_data["actualStart"]["value"]
            date_obj = datetime.strptime(raw_date, "%Y-%m-%dT%H:%M:%S%z")
            formatted_date = f"{date_obj.day} {MONTHS_RU[date_obj.month]}"

            for product in products:
                if is_product(product):
                    sheet.cell(row=row, column=1, value=row - 1).alignment = Alignment(horizontal="center",
                                                                                       vertical="center")
                    sheet.cell(row=row, column=2, value=project_name).alignment = Alignment(horizontal="center",
                                                                                            vertical="center")
                    sheet.cell(row=row, column=3, value=issue_name).alignment = Alignment(horizontal="center",
                                                                                          vertical="center")

                    sheet.cell(row=row, column=4, value=product).alignment = Alignment(wrap_text=True,
                                                                                       horizontal="left",
                                                                                       vertical="center")

                    sheet.cell(row=row, column=5, value=formatted_date).alignment = Alignment(horizontal="center",
                                                                                              vertical="center")

                    sheet.cell(row=row, column=6, value=project_responsible).alignment = Alignment(horizontal="center",
                                                                                                   vertical="center")
                    sheet.cell(row=row, column=7, value=owner_name).alignment = Alignment(
                        horizontal="center", vertical="center")

                    sheet.cell(row=row, column=8, value=raw_materials_comment).alignment = Alignment(
                        horizontal="center",
                        vertical="center")
                    sheet.cell(row=row, column=9, value=packaging_comment).alignment = Alignment(horizontal="center",
                                                                                                 vertical="center")
                    sheet.cell(row=row, column=10, value=last_comment).alignment = Alignment(horizontal="center",
                                                                                             vertical="center")

                    row += 1


def upload_file(filename, real_name):
    url = f"{MEGAPLAN_API_URL}/api/file"
    headers = {
        "Authorization": f"Bearer {MEGAPLAN_API_KEY}"

    }
    real_name = real_name = f"{real_name}.{filename.rsplit('.', 1)[-1]}"
    with open(filename, 'rb') as file:
        files = {'files[]': (real_name, file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        try:
            response = requests.post(url, headers=headers, files=files)
            response.raise_for_status()
            file_data = response.json()['data'][0]
            file_id = file_data['id']
            return file_id
        except requests.RequestException as e:
            logging.exception(f"Error uploading file: {e}")
            return None
        finally:
            os.remove(filename)


@router.get("/app/test")
async def test_endpoint():
    return JSONResponse(status_code=200, content={"message": "Test request successful!"})


class EntityRequest(BaseModel):
    entityType: str
    entityId: str


@router.post("/app/unloading-tasks")
async def unload_tasks(request: EntityRequest):
    logging.info(f"Webhook_data: {request.json()}")
    entity_type = request.entityType.lower()
    entity_id = request.entityId

    if entity_type not in ["project", "task"]:
        raise HTTPException(status_code=400, detail="Invalid entityType. Must be 'project' or 'task'.")

    # Создаем асинхронную задачу для обработки выгрузки
    asyncio.create_task(process_tasks_unloading(entity_type, entity_id))
    return JSONResponse(status_code=200, content={"message": "Задача выгрузки принята в обработку"})


async def process_tasks_unloading(entity_type: str, entity_id: str):
    # Определяем URL для комментария и структуру subject в зависимости от типа сущности
    if entity_type == "project":
        comment_url = f"{MEGAPLAN_API_URL}/api/v3/project/{entity_id}/comments"
        subject = {
            "id": entity_id,
            "contentType": "Project"
        }
    elif entity_type == "task":
        comment_url = f"{MEGAPLAN_API_URL}/api/v3/task/{entity_id}/comments"
        subject = {
            "id": entity_id,
            "contentType": "Task"
        }
    else:
        # Эта проверка уже сделана в маршруте, но оставляем её для дополнительной безопасности
        raise HTTPException(status_code=400, detail="Unsupported entityType")

    try:
        # Создаем временный файл Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Задачи"

            # Настройка стилей для первой строки
            header_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")  # Жёлтый цвет
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Границы для разграничения заголовков
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Установка заголовков и применение стиля ко всей первой строке
            headers = ["№", "Бренд", "Линейка", "Наименование", "Дата запуска работы", "Ответственный БМ",
                       "Ответственный ОЗ", "Сырье", "Упаковка", "Примечание"]
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill  # Применяем жёлтый цвет
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border  # Добавляем границы

            # Получение данных в зависимости от типа сущности
            if entity_type == "project":
                issues = get_project_issues(entity_id, MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                project_data = get_project(entity_id, MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                project_name = project_data["name"]
                project_responsible = get_responsible_name(project_data["responsible"], MEGAPLAN_API_URL,
                                                           MEGAPLAN_HEADER)
            elif entity_type == "task":
                task_data = get_task(entity_id, MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                issues = get_task_subtasks(entity_id, MEGAPLAN_API_URL, MEGAPLAN_HEADER)
                project_name = task_data["name"]
                project_responsible = get_responsible_name(task_data["responsible"], MEGAPLAN_API_URL,
                                                           MEGAPLAN_HEADER)
            else:
                # Эта проверка уже сделана, но оставляем её для дополнительной безопасности
                raise HTTPException(status_code=400, detail="Unsupported entityType")

            # Запуск обработки задач
            process_tasks(project_name, issues, sheet, project_responsible)

            # Настройка ширины колонок
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells) + 2  # Добавляем дополнительную ширину
                sheet.column_dimensions[column_cells[0].column_letter].width = min(length,
                                                                                   50)  # Ограничение максимальной шириной

            sheet.column_dimensions['A'].width = 5
            sheet.column_dimensions['D'].width = 50

            # Ограничение ширины столбцов сырье/упаковка/примечание
            max_width = 50  # Максимальная ширина в символах
            sheet.column_dimensions['H'].width = max_width  # Столбец "Сырье"
            sheet.column_dimensions['I'].width = max_width  # Столбец "Упаковка"
            sheet.column_dimensions['J'].width = max_width  # Столбец "Примечание"

            # Перенос текста для столбцов 'H', 'I', 'J'
            for column_letter in ['H', 'I', 'J']:
                for cell in sheet[column_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

            workbook.save(tmp.name)
            tmp.close()

        # Загружаем файл
        file_id = upload_file(tmp.name, real_name=project_name)

        if not file_id:
            raise HTTPException(status_code=500, detail="Error uploading file")

        # Отправляем комментарий с файлом
        content_text = f"Задачи проекта {project_name}"
        body = {
            "contentType": "CommentCreateActionRequest",
            "comment": {
                "contentType": "Comment",
                "content": content_text,
                "attaches": [
                    {
                        "id": file_id,
                        "contentType": "File"
                    }
                ],
                "subject": subject
            },
            "transports": [
                {}
            ]
        }

        try:
            response = requests.post(comment_url, headers=MEGAPLAN_HEADER, json=body)
            response.raise_for_status()
            logging.info(
                f"Комментарий успешно отправлен для {'проекта' if entity_type == 'project' else 'задачи'} с ID: {entity_id}")
        except requests.RequestException as e:
            logging.exception(f"Error posting comment: {e}")
            raise HTTPException(status_code=500, detail="Error posting comment")

    except Exception as e:
        logging.exception(f"Error during process_tasks_unloading: {e}")
        # Отправляем комментарий с сообщением об ошибке
        error_content = "[KUBIT — Отчет] Ошибка структуры данных. Проверьте соблюдение иерархии"
        error_body = {
            "contentType": "CommentCreateActionRequest",
            "comment": {
                "contentType": "Comment",
                "content": error_content,
                "attaches": [],
                "subject": subject
            },
            "transports": [
                {}
            ]
        }
        try:
            error_response = requests.post(comment_url, headers=MEGAPLAN_HEADER, json=error_body)
            error_response.raise_for_status()
            logging.info(
                f"Комментарий об ошибке успешно отправлен для {'проекта' if entity_type == 'project' else 'задачи'} с ID: {entity_id}")
        except requests.RequestException as ex:
            logging.exception(f"Error sending error comment: {ex}")
